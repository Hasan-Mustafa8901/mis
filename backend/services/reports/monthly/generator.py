# backend\services\reports\monthly\generator.py

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

TITLE_FONT = Font(name="Times New Roman", bold=True, size=16)
HEADER_FONT = Font(name="Times New Roman", bold=True, size=11)
BODY_FONT = Font(name="Times New Roman", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")


def _setup_columns(ws):

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def _write_header(ws, row, report):

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1)
    cell.value = "MONTHLY STATISTICS REPORT"
    cell.font = TITLE_FONT
    row += 2
    ws.cell(row=row, column=1, value="DEALERSHIP").font = HEADER_FONT
    ws.cell(row=row, column=2, value=report.dealership_name.upper())
    row += 1
    ws.cell(row=row, column=1, value="PERIOD").font = HEADER_FONT
    ws.cell(
        row=row,
        column=2,
        value=(f"{report.report_period_from} TO {report.report_period_to}"),
    )

    return row + 2


def _section_header(ws, row, title):

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1)
    cell.value = title.upper()
    cell.font = HEADER_FONT
    cell.fill = SECTION_FILL

    return row + 1


def _write_reconciliation(ws, row, report):

    row = _section_header(ws, row, "Reconciliation")
    metrics = [
        ("TOTAL VEHICLE BOOKED", report.total_vehicle_booked),
        ("TOTAL VEHICLE DELIVERED [A]", report.total_vehicle_delivered),
        ("TOTAL OUT OF AUDIT PURVIEW [B]", report.total_out_of_audit_purview),
        (
            "TOTAL DELIVERY CASES TO BE VERIFIED [C]",
            report.total_delivery_cases_to_be_verified,
        ),
        ("FILES PENDING VERIFICATION [D]", report.files_pending_verification),
        ("TOTAL DELIVERY CASES VERIFIED [E]", report.total_delivery_cases_verified),
    ]

    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

        row += 1

    return row + 1


def _write_category_discount(ws, row, report):

    row = _section_header(ws, row, "Category Wise Discount")
    ws.cell(row=row, column=1, value="DISCOUNT COMPONENT").font = Font(bold=True)
    ws.cell(row=row, column=2, value="AMOUNT").font = Font(bold=True)
    row += 1

    for item in report.category_discounts:
        ws.cell(row=row, column=1, value=item.component.upper())
        ws.cell(row=row, column=2, value=item.amount)
        row += 1

    return row + 1


def _write_discount_summary(ws, row, report):

    row = _section_header(ws, row, "Discount Summary")

    metrics = [
        ("Total Discount Given", report.total_discount_given),
        ("Maximum Allowable Discount", report.maximum_allowable_discount),
        ("Excess Discount Given", report.excess_discount_given),
        ("Average Discount", report.average_discount),
        ("Average Excess Discount", report.average_excess_discount),
    ]

    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    return row + 1


def _write_model_analysis(ws, row, report):
    row = _section_header(ws, row, "Model / Fuel Type Analysis")
    grouped = {}
    for item in report.model_discount_analysis:
        grouped.setdefault(item.car_name, []).append(item)

    for model_name, rows in grouped.items():
        ws.cell(row=row, column=1, value=f"MODEL : {model_name.upper()}").font = Font(
            bold=True
        )

        row += 1

        headers = [
            "FUEL TYPE",
            "DELIVERED",
            "TOTAL DISCOUNT",
            "AVG DISCOUNT",
            "EXCESS DISCOUNT",
            "AVG EXCESS",
        ]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)

        row += 1

        for item in rows:
            ws.cell(row=row, column=1, value=item.fuel_type.upper())
            ws.cell(row=row, column=2, value=item.delivered_cases)
            ws.cell(row=row, column=3, value=item.total_discount)
            ws.cell(row=row, column=4, value=item.average_discount)
            ws.cell(row=row, column=5, value=item.total_excess_discount)
            ws.cell(row=row, column=6, value=item.average_excess_discount)

            row += 1

        row += 1

    return row


def _write_showroom_analysis(ws, row, report):

    row = _section_header(ws, row, "Showroom Wise Model Analysis")

    if not report.showroom_model_analysis:
        ws.cell(row=row, column=1, value="NO DATA AVAILABLE")
        return row + 2

    outlets = sorted({r.outlet_name for r in report.showroom_model_analysis})

    pivot = {}

    for r in report.showroom_model_analysis:
        key = (r.car_name, r.fuel_type)
        pivot.setdefault(key, {})[r.outlet_name] = r

    # HEADER ROW 1

    ws.cell(row=row, column=1, value="MODEL / FUEL").font = HEADER_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    start_col = 2

    for outlet in outlets + ["TOTAL"]:
        end_col = start_col + 4

        ws.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=end_col
        )

        cell = ws.cell(row=row, column=start_col, value=outlet.upper())

        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

        start_col += 5

    row += 1

    # HEADER ROW 2

    metrics = [
        "No of Vehicles",
        "Discount Given",
        "Average Discount Given",
        "Excess Discount Given",
        "Average Excess Discount",
    ]

    col = 2
    for _ in outlets + ["TOTAL"]:
        for metric in metrics:
            cell = ws.cell(row=row, column=col, value=metric)

            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

            col += 1

    row += 1

    # DATA ROWS
    grand_totals = {
        outlet: {
            "vehicles": 0,
            "discount": 0,
            "excess": 0,
        }
        for outlet in outlets
    }

    for car_name, fuel_type in sorted(pivot.keys()):
        ws.cell(
            row=row, column=1, value=f"{car_name} {fuel_type}".upper()
        ).font = BODY_FONT

        col = 2

        total_vehicles = 0
        total_discount = 0
        total_excess = 0

        for outlet in outlets:
            data = pivot[(car_name, fuel_type)].get(outlet)

            if data:
                vehicles = data.delivered_cases
                discount = data.total_discount
                avg_discount = data.average_discount
                excess = data.total_excess_discount
                avg_excess = data.average_excess_discount

            else:
                vehicles = 0
                discount = 0
                avg_discount = 0
                excess = 0
                avg_excess = 0

            total_vehicles += vehicles
            total_discount += discount
            total_excess += excess

            grand_totals[outlet]["vehicles"] += vehicles

            grand_totals[outlet]["discount"] += discount

            grand_totals[outlet]["excess"] += excess

            values = [
                vehicles,
                round(discount, 2),
                round(avg_discount, 2),
                round(excess, 2),
                round(avg_excess, 2),
            ]

            for value in values:
                cell = ws.cell(row=row, column=col, value=value)

                cell.font = BODY_FONT
                cell.border = THIN_BORDER

                col += 1

        # TOTAL BLOCK
        overall_avg_discount = total_discount / total_vehicles if total_vehicles else 0

        overall_avg_excess = total_excess / total_vehicles if total_vehicles else 0

        totals = [
            total_vehicles,
            round(total_discount, 2),
            round(overall_avg_discount, 2),
            round(total_excess, 2),
            round(overall_avg_excess, 2),
        ]

        for value in totals:
            cell = ws.cell(row=row, column=col, value=value)

            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            col += 1

        row += 1

    # GRAND TOTAL ROW

    ws.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT

    col = 2

    overall_vehicle_total = 0
    overall_discount_total = 0
    overall_excess_total = 0

    for outlet in outlets:
        vehicles = grand_totals[outlet]["vehicles"]

        discount = grand_totals[outlet]["discount"]

        excess = grand_totals[outlet]["excess"]

        avg_discount = discount / vehicles if vehicles else 0

        avg_excess = excess / vehicles if vehicles else 0

        overall_vehicle_total += vehicles
        overall_discount_total += discount
        overall_excess_total += excess

        values = [
            vehicles,
            round(discount, 2),
            round(avg_discount, 2),
            round(excess, 2),
            round(avg_excess, 2),
        ]

        for value in values:
            cell = ws.cell(row=row, column=col, value=value)

            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

            col += 1

    overall_avg_discount = (
        overall_discount_total / overall_vehicle_total if overall_vehicle_total else 0
    )

    overall_avg_excess = (
        overall_excess_total / overall_vehicle_total if overall_vehicle_total else 0
    )

    totals = [
        overall_vehicle_total,
        round(overall_discount_total, 2),
        round(overall_avg_discount, 2),
        round(overall_excess_total, 2),
        round(overall_avg_excess, 2),
    ]

    for value in totals:
        cell = ws.cell(row=row, column=col, value=value)

        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

        col += 1

    return row + 2


def generate_monthly_report(report):
    wb = Workbook()

    ws = wb.active
    ws.title = "Monthly Statistics"

    _setup_columns(ws)

    row = 1

    row = _write_header(ws, row, report)
    row = _write_reconciliation(ws, row, report)
    row = _write_category_discount(ws, row, report)
    row = _write_discount_summary(ws, row, report)
    row = _write_model_analysis(ws, row, report)
    row = _write_showroom_analysis(ws, row, report)

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    filename = (
        f"{report.dealership_name.title()} Audit Report"
        f"{report.report_period_from.replace('/', '-')}"
        f"_to_"
        f"{report.report_period_to.replace('/', '-')}"
        f".xlsx"
    )

    return buffer, filename
