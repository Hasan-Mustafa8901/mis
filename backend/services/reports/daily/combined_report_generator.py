# backend\services\reports\daily\combined_report_generator.py.py
"""
report_generator.py
────────────────────────────────────────────────────────────────────────────────
Excel Report Generator — Combined (N-Dealer)
────────────────────────────────────────────────────────────────────────────────
PUBLIC API
──────────
    generate_report(backend_data)  ← single entry point for frontend

BACKEND_DATA SCHEMAS

    {
        "report_date": "30/04/2026",        # same formats as daily
        "scope": "All Dealers",             # label shown in title

        "dealers": [                        # 1 or more dealers — columns auto-expand
            {
                "name":     "Dealer Name 1",
                "booking":  { <METRIC_KEYS> },
                "delivery": { <METRIC_KEYS> },
            },
            {
                "name":     "Dealer Name 2",
                "booking":  { <METRIC_KEYS> },
                "delivery": { <METRIC_KEYS> },
            },
            # ...add as many dealers as needed
        ],
    }

── METRIC_KEYS (used by both report types) ─────────────────────────────────────

    # Files Reconciliation table
    "Total Cases Reported"                              : int
    "Files Received"                                    : int
    "Files Pending"                                     : int          ← red highlight
    "Files Incomplete"                                  : int          ← red highlight
    "Files Verified"                                    : int
    "Files Approved"                                    : int
    "Files Rejected"                                    : int          ← red highlight
    "Verification Completion %"                         : float 0-1    ← shown as 0%..100%

    # Discount Summary table
    "Total Discount Given"                              : int/float    ← Indian comma format
    "Discount as per Approved Scheme"                   : int/float    ← Indian comma format
    "Net Excess Discount Amount"                        : int/float    ← Indian comma format
    "Highest Discount Car Model"                        : str
    "Highest Discount Value"                            : int/float    ← Indian comma format
    "Excess Discount Cases"                             : int
    "Allowable Discount Cases (out of Verified cases)"  : int
    "Excess Discount Cases(out of Verified cases)"      : int          ← red highlight
    "Zero Discount Cases(out of Verified cases)"        : int
"""

from io import BytesIO
from datetime import date, datetime
from typing import cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# # Keys whose values get Indian comma number format
_MONEY_KEYWORDS = ("Amount", "Value", "Discount Given", "Approved Scheme")

# Rows that are red-highlighted in the Files Reconciliation table
_RECON_HIGHLIGHT = {"Files Pending", "Files Incomplete", "Files Rejected"}

# Rows that are red-highlighted in the Discount Summary table
_DISCOUNT_HIGHLIGHT = {"Excess Discount Cases(out of Verified cases)"}

# Ordered row definitions for both tables
_RECON_ROWS = [
    "Total Cases Reported",
    "Files Received",
    "Files Pending",
    "Files Out of Scope",
    "Files Incomplete",
    "Files Verified",
    "Files Approved",
    "Files Rejected",
    "Verification Completion %",
]

_DISCOUNT_ROWS = [
    "Total Discount Given",
    "Discount as per Approved Scheme",
    "Net Excess Discount Amount",
    "Highest Discount Car Model",
    "Highest Discount Value",
    "Excess Discount Cases",
    "Allowable Discount Cases (out of Verified cases)",
    "Excess Discount Cases(out of Verified cases)",
    "Zero Discount Cases(out of Verified cases)",
]


def _styles() -> dict:
    """Return a shared set of openpyxl style objects."""
    return dict(
        font_bold=Font(name="Times New Roman", bold=True, color="000000"),
        font_head=Font(name="Times New Roman", bold=True, size=15, color="000000"),
        font_normal=Font(name="Times New Roman", bold=False, color="000000"),
        align_center=Alignment(horizontal="center", vertical="center", wrap_text=True),
        align_left=Alignment(horizontal="left", vertical="center", wrap_text=True),
        align_right=Alignment(horizontal="right", vertical="center"),
        med=Side(style="medium"),
        thn=Side(style="thin"),
        fill_grey=PatternFill("solid", fgColor="E7E6E6"),
        fill_blue_hdr=PatternFill("solid", fgColor="9BC2E6"),
        fill_blue_sub=PatternFill("solid", fgColor="D9E1F2"),
        fill_red=PatternFill("solid", fgColor="FF6D6D"),
        fill_none=PatternFill(fill_type=None),
    )


# SHARED INTERNALS
def _col_letter(n: int) -> str:
    """1-based column index → Excel column letter. e.g. 1→A, 27→AA."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


class CombinedReportGenerator:
    def __init__(self, backend_data: dict) -> None:
        self.backend_data = backend_data or {}

        self.report_scope = self.backend_data.get("scope", "All Dealers")
        self.date_str = self._resolve_date_str(self.backend_data.get("report_date", ""))
        self.dealers = self.backend_data.get("dealers", [])

        if not self.dealers:
            self.dealers = [
                {
                    "name": "Dealer 1",
                    "booking": {},
                    "delivery": {},
                }
            ]
        self.num_dealers = len(self.dealers)
        self.num_cols = 3 + self.num_dealers * 2
        self.last_col = _col_letter(self.num_cols)

        self.wb: Workbook = Workbook()
        self.ws: Worksheet = cast(Worksheet, self.wb.active)
        self.ws.title = "Combined Report"

        S = _styles()
        self.font_bold = S["font_bold"]
        self.font_head = S["font_head"]
        self.font_normal = S["font_normal"]
        self.align_center = S["align_center"]
        self.align_left = S["align_left"]
        self.align_right = S["align_right"]
        self.med = S["med"]
        self.thn = S["thn"]
        self.fill_grey = S["fill_grey"]
        self.fill_blue_hdr = S["fill_blue_hdr"]
        self.fill_blue_sub = S["fill_blue_sub"]
        self.fill_red = S["fill_red"]

    def generate(self):
        self._configure_columns()
        self._render_title()

        self._render_metric_table(
            title="Files Reconciliation",
            rows=_RECON_ROWS,
            highlight_rows=_RECON_HIGHLIGHT,
            section_row=4,
            header_row=5,
            data_start_row=7,
        )
        self._render_metric_table(
            title="Discount Summary",
            rows=_DISCOUNT_ROWS,
            highlight_rows=_DISCOUNT_HIGHLIGHT,
            section_row=16,
            header_row=17,
            data_start_row=19,
        )
        return self._save()

    def _configure_columns(self):
        self.ws.column_dimensions["A"].width = 6
        self.ws.column_dimensions["B"].width = 12
        self.ws.column_dimensions["C"].width = 33

        for idx in range(self.num_dealers * 2):
            self.ws.column_dimensions[_col_letter(4 + idx)].width = 18

    def _render_title(self, row_start: int = 1):
        self.ws.merge_cells(f"A{row_start}:{self.last_col}{row_start + 1}")

        cell = self.ws.cell(
            row=row_start,
            column=1,
            value=f"Report for {self.report_scope}\nAs on ({self.date_str})",
        )

        cell.font = self.font_head
        cell.alignment = self.align_center
        cell.fill = self.fill_grey

        for row in range(row_start, row_start + 2):
            self.ws.row_dimensions[row].height = 20

            for col in range(1, self.num_cols + 1):
                self.ws.cell(row=row, column=col).border = Border(
                    left=self.med if col == 1 else Side(style=None),
                    right=self.med if col == self.num_cols else Side(style=None),
                    top=self.med if row == row_start else Side(style=None),
                    bottom=self.med if row == row_start + 1 else Side(style=None),
                )

    def _render_section_header(self, row: int, title: str):
        self.ws.merge_cells(f"A{row}:{self.last_col}{row}")

        cell = self.ws.cell(
            row=row,
            column=1,
            value=title,
        )

        cell.font = self.font_bold
        cell.alignment = self.align_center
        cell.fill = self.fill_blue_hdr

        self.ws.row_dimensions[row].height = 18

        for col in range(1, self.num_cols + 1):
            self.ws.cell(row=row, column=col).border = Border(
                left=self.med if col == 1 else Side(style=None),
                right=self.med if col == self.num_cols else Side(style=None),
                top=self.med,
                bottom=self.med,
            )

    def _render_column_headers(self, header_row: int):
        self.ws.merge_cells(f"A{header_row}:C{header_row + 1}")

        cell = self.ws.cell(row=header_row, column=1, value="Particulars")

        cell.font = self.font_bold
        cell.alignment = self.align_center
        cell.fill = self.fill_blue_sub

        for row in range(header_row, header_row + 2):
            for col in range(1, 4):
                self.ws.cell(row=row, column=col).fill = self.fill_blue_sub

                self.ws.cell(row=row, column=col).border = Border(
                    left=self.med if col == 1 else self.thn,
                    right=self.thn,
                    top=self.med if row == header_row else self.thn,
                    bottom=self.thn,
                )

        for dealer_index, dealer in enumerate(self.dealers):
            booking_col = 4 + dealer_index * 2
            delivery_col = booking_col + 1

            booking_letter = _col_letter(booking_col)
            delivery_letter = _col_letter(delivery_col)

            self.ws.merge_cells(
                f"{booking_letter}{header_row}:{delivery_letter}{header_row}"
            )

            dealer_cell = self.ws.cell(
                row=header_row, column=booking_col, value=dealer["name"]
            )

            dealer_cell.font = self.font_bold
            dealer_cell.alignment = self.align_center
            dealer_cell.fill = self.fill_blue_sub

            for col in [booking_col, delivery_col]:
                self.ws.cell(row=header_row, column=col).fill = self.fill_blue_sub

                self.ws.cell(row=header_row, column=col).border = Border(
                    left=self.thn,
                    right=self.med if col == self.num_cols else self.thn,
                    top=self.med,
                    bottom=self.thn,
                )

            for col, label in ((booking_col, "Booking"), (delivery_col, "Delivery")):
                cell = self.ws.cell(row=header_row + 1, column=col, value=label)

                cell.font = self.font_bold
                cell.alignment = self.align_center
                cell.fill = self.fill_blue_sub

                cell.border = Border(
                    left=self.thn,
                    right=self.med if col == self.num_cols else self.thn,
                    top=self.thn,
                    bottom=self.thn,
                )

        self.ws.row_dimensions[header_row].height = 18
        self.ws.row_dimensions[header_row + 1].height = 16

    def _get_metric_values(self, metric_name: str) -> list:
        values = []

        for dealer in self.dealers:
            values.extend(
                [
                    dealer.get("booking", {}).get(metric_name, ""),
                    dealer.get("delivery", {}).get(metric_name, ""),
                ]
            )

        return values

    def _render_metric_row(
        self,
        row: int,
        label: str,
        values: list,
        highlight: bool = False,
        is_last: bool = False,
    ):
        bottom = self.med if is_last else self.thn
        fill = self.fill_red if highlight else PatternFill(fill_type=None)

        self.ws.row_dimensions[row].height = 15

        self.ws.merge_cells(f"A{row}:C{row}")

        label_cell = self.ws.cell(row=row, column=1, value=label)

        label_cell.font = self.font_normal
        label_cell.alignment = self.align_left
        label_cell.fill = fill

        label_cell.border = Border(
            left=self.med,
            right=self.thn,
            top=self.thn,
            bottom=bottom,
        )

        for col in [2, 3]:
            self.ws.cell(row=row, column=col).fill = fill

            self.ws.cell(row=row, column=col).border = Border(
                left=self.thn,
                right=self.thn,
                top=self.thn,
                bottom=bottom,
            )

        data_cells = []

        for idx, value in enumerate(values):
            col = 4 + idx

            cell = self.ws.cell(row=row, column=col, value=value)

            cell.font = self.font_normal
            cell.fill = fill

            cell.alignment = (
                self.align_right
                if isinstance(value, (int, float))
                else self.align_center
            )

            cell.border = Border(
                left=self.thn,
                right=self.med if col == self.num_cols else self.thn,
                top=self.thn,
                bottom=bottom,
            )

            data_cells.append(cell)

        return data_cells

    def _render_metric_table(
        self,
        title: str,
        rows: list[str],
        highlight_rows: set[str],
        section_row: int,
        header_row: int,
        data_start_row: int,
    ):
        self._render_section_header(section_row, title)

        self._render_column_headers(header_row)

        for index, metric in enumerate(rows):
            row = data_start_row + index

            data_cells = self._render_metric_row(
                row=row,
                label=metric,
                values=self._get_metric_values(metric),
                highlight=metric in highlight_rows,
                is_last=index == len(rows) - 1,
            )

            if metric == "Verification Completion %":
                for cell in data_cells:
                    cell.number_format = "0%"

            if any(keyword in metric for keyword in _MONEY_KEYWORDS):
                for cell in data_cells:
                    cell.number_format = "##,##,##0"

    def _save(self):
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        safe_date = self.date_str.replace("/", "-") if self.date_str else "no-date"
        return (buffer, f"Combined-Report-{safe_date}.xlsx")

    def _resolve_date_str(self, raw) -> str:
        """Normalise any date input to a display string."""
        if isinstance(raw, (date, datetime)):
            return raw.strftime("%d/%m/%Y")

        if isinstance(raw, dict):
            frm = raw.get("from")
            to = raw.get("to")

            if isinstance(frm, (date, datetime)):
                frm = frm.strftime("%d/%m/%Y")

            if isinstance(to, (date, datetime)):
                to = to.strftime("%d/%m/%Y")

            if frm and to:
                return f"{frm} To {to}"

            return str(frm or to or "")

        return str(raw) if raw else ""


if __name__ == "__main__":
    import os

    METRIC = {
        "Total Cases Reported": 100,
        "Files Received": 95,
        "Files Pending": 7,
        "Files Incomplete": 2,
        "Files Verified": 86,
        "Files Approved": 80,
        "Files Rejected": 6,
        "Verification Completion %": 0.86,
        "Total Discount Given": 450000,
        "Discount as per Approved Scheme": 400000,
        "Net Excess Discount Amount": 50000,
        "Highest Discount Car Model": "Swift Dzire",
        "Highest Discount Value": 30000,
        "Excess Discount Cases": 10,
        "Allowable Discount Cases (out of Verified cases)": 76,
        "Excess Discount Cases(out of Verified cases)": 10,
        "Zero Discount Cases(out of Verified cases)": 3,
    }

    def scaled(n):
        return {
            k: (
                round(v * n, 2)
                if isinstance(v, float)
                else (f"{v} {n}" if isinstance(v, str) else int(v * n))
            )
            for k, v in METRIC.items()
        }

    test_cases = [
        {
            "report_date": "22/05/2026",
            "scope": "Single Dealer",
            "dealers": [
                {
                    "name": "Dealer Alpha",
                    "booking": scaled(1),
                    "delivery": scaled(0.8),
                }
            ],
        },
        {
            "report_date": "22/05/2026",
            "scope": "All Dealers",
            "dealers": [
                {
                    "name": "Dealer Alpha",
                    "booking": scaled(1),
                    "delivery": scaled(0.8),
                },
                {
                    "name": "Dealer Beta",
                    "booking": scaled(0.7),
                    "delivery": scaled(0.6),
                },
            ],
        },
        {
            "report_date": "22/05/2026",
            "scope": "All Dealers",
            "dealers": [
                {
                    "name": "Dealer Alpha",
                    "booking": scaled(1),
                    "delivery": scaled(0.8),
                },
                {
                    "name": "Dealer Beta",
                    "booking": scaled(0.7),
                    "delivery": scaled(0.6),
                },
                {
                    "name": "Dealer Gamma",
                    "booking": scaled(0.5),
                    "delivery": scaled(0.4),
                },
            ],
        },
    ]

    output_dir = "backend/test_reports"
    os.makedirs(output_dir, exist_ok=True)

    for idx, data in enumerate(test_cases, start=1):
        # Option 2
        buf, fname = CombinedReportGenerator(data).generate()

        path = os.path.join(output_dir, f"test_{idx}_{fname}")

        with open(path, "wb") as f:
            f.write(buf.getvalue())

        print(f"✓ Generated: {path} (dealers={len(data['dealers'])})")

    print("\nAll test reports generated successfully.")
