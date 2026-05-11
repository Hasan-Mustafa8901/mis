from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import date, datetime


def generate_daily_report(backend_data=None):
    """
    Generate a formatted Excel daily report and return (BytesIO, filename).

    All data is supplied via backend_data. Structure expected:

    backend_data = {
        "report_date": "30/04/2026",          # str dd/mm/yyyy  OR  date/datetime object
        OR
        "start_date" and "end_date

        "booking": {
            "Total Cases Reported":                             <int>,
            "Files Received":                                   <int>,
            "Files Pending":                                    <int>,
            "Files Incomplete":                                 <int>,
            "Files Verified":                                   <int>,
            "Files Approved":                                   <int>,
            "Files Rejected":                                   <int>,
            "Verification Completion %":                        <float 0-1>,
            "Total Discount Given":                             <int/float>,
            "Discount as per Approved Scheme":                  <int/float>,
            "Net Excess Discount Amount":                       <int/float>,
            "Highest Discount Car Model":                       <str>,
            "Highest Discount Value":                           <int/float>,
            "Excess Discount Cases":                            <int>,
            "Allowable Discount Cases (out of Verified cases)": <int>,
            "Excess Discount Cases(out of Verified cases)":     <int>,
            "Zero Discount Cases(out of Verified cases)":       <int>,
        },

        "delivery": {
            # same keys as "booking" above
        },

        # ── Sheet 1: Pending Files ─────────────────────────────────────────
        "booking_files_pending": [
            # one dict per row, up to 10 rows displayed
            {
                "sno":    <int>,   # Serial number
                "date":   <str>,   # e.g. "26/01/2026"
                "name":   <str>,   # Customer name
                "mobile": <str>,   # Mobile number
                "tl":     <str>,   # Team Leader name
            },
            ...
        ],

        "delivery_files_pending": [
            # same structure as booking_files_pending
        ],

        # ── Sheet 2: Pending Documents ─────────────────────────────────────
        "booking_docs_pending": [
            # one dict per row, up to 10 rows displayed
            {
                "sno":                <int>,
                "date":               <str>,
                "name":               <str>,
                "mobile":             <str>,
                "tl":                 <str>,
                # Booking checklist items — values: "Received", "Pending", "Partial", "NA"
                "kyc":                <str>,   # Customer KYC
                "vehicle":            <str>,   # Vehicle Details
                "quotation":          <str>,   # Price Quotation
                "receipts":           <str>,   # Receipts
                "accessories_indent": <str>,   # Accessories Indent
                "exchange":           <str>,   # Exchange Details
                "md_approval":        <str>,   # MD Reference Approval
                "corp_id":            <str>,   # Corp ID
                "customer_sign":      <str>,   # Customer Sign
            },
            ...
        ],

        "delivery_docs_pending": [
            # one dict per row, up to 10 rows displayed
            {
                "sno":                <int>,
                "date":               <str>,
                "name":               <str>,
                "mobile":             <str>,
                "tl":                 <str>,
                # Delivery checklist items — values: "Received", "Pending", "Partial", "NA"
                "ledger":             <str>,   # Customer Ledger
                "tax_invoice":        <str>,   # Tax Invoice
                "accessories_indent": <str>,   # Accessories Indent
                "insurance":          <str>,   # Insurance
                "rto":                <str>,   # RTO
                "finance":            <str>,   # Finance
                "eval_cert":          <str>,   # Evaluation Certificate
            },
            ...
        ],
    }
    """
    if backend_data is None:
        backend_data = {}

    # ── Resolve report date ───────────────────────────────────────────────────
    raw_date = backend_data.get("report_date", "")

    # ── Single Date ───────────────────────────────────────────
    if isinstance(raw_date, (date, datetime)):
        date_str = raw_date.strftime("%d/%m/%Y")

    # ── Date Interval ─────────────────────────────────────────
    elif isinstance(raw_date, dict):
        from_date = raw_date.get("from")
        to_date = raw_date.get("to")

        if isinstance(from_date, (date, datetime)):
            from_date = from_date.strftime("%d/%m/%Y")
        if isinstance(to_date, (date, datetime)):
            to_date = to_date.strftime("%d/%m/%Y")
        if from_date and to_date:
            date_str = f"{from_date} To {to_date}"
        elif from_date:
            date_str = str(from_date)
        elif to_date:
            date_str = str(to_date)
        else:
            date_str = ""

    # ── Plain String ──────────────────────────────────────────
    elif raw_date:
        date_str = str(raw_date)
    else:
        date_str = ""

    booking = backend_data.get("booking", {})
    delivery = backend_data.get("delivery", {})

    # Sheet 1 pending files (new split structure)
    booking_files_pending = backend_data.get("booking_files_pending", [])
    delivery_files_pending = backend_data.get("delivery_files_pending", [])

    # Out of Scope files
    booking_out_of_scope = backend_data.get(
        "booking_out_of_scope",
        [],
    )

    delivery_out_of_scope = backend_data.get(
        "delivery_out_of_scope",
        [],
    )
    booking_delay_files = backend_data.get(
        "booking_delay_files",
        [],
    )

    delivery_delay_files = backend_data.get(
        "delivery_delay_files",
        [],
    )
    rejected_files_delivered = backend_data.get(
        "rejected_files_delivered",
        [],
    )
    # Sheet 2 pending docs
    booking_docs_pending = backend_data.get("booking_docs_pending", [])
    delivery_docs_pending = backend_data.get("delivery_docs_pending", [])

    # ── Workbook / sheet ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    # ── Styles ────────────────────────────────────────────────────────────────
    font_bold = Font(name="Times New Roman", bold=True, color="000000")
    font_head = Font(name="Times New Roman", bold=True, size=15, color="000000")
    font_normal = Font(name="Times New Roman", color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    med = Side(style="medium")
    thn = Side(style="thin")

    fill_grey = PatternFill("solid", fgColor="E7E6E6")
    fill_blue_hdr = PatternFill("solid", fgColor="9BC2E6")
    fill_blue_sub = PatternFill("solid", fgColor="D9E1F2")
    fill_red = PatternFill("solid", fgColor="FF6D6D")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 33
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # ── Helpers ───────────────────────────────────────────────────────────────
    def write_title(row_start):
        ws.merge_cells(f"A{row_start}:E{row_start + 1}")
        c = ws.cell(row=row_start, column=1, value=f"Daily Report As on ({date_str})")
        c.font = font_head
        c.alignment = align_center
        c.fill = fill_grey
        for r in range(row_start, row_start + 2):
            ws.row_dimensions[r].height = 20
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else Side(style=None),
                    right=med if col == 5 else Side(style=None),
                    top=med if r == row_start else Side(style=None),
                    bottom=med if r == row_start + 1 else Side(style=None),
                )

    def write_section_header(row, title):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_hdr
        ws.row_dimensions[row].height = 18
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(
                left=med if col == 1 else Side(style=None),
                right=med if col == 5 else Side(style=None),
                top=med,
                bottom=med,
            )

    def write_col_headers_merged(row, d_label="Booking", e_label="Delivery"):
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 16
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == 5 else thn,
                top=thn,
                bottom=thn,
            )
        ws.cell(row=row, column=1).value = "Particulars"
        ws.cell(row=row, column=4).value = d_label
        ws.cell(row=row, column=5).value = e_label

    def write_data_row(row, label, b_val, d_val, highlight=False, is_last=False):
        bottom = med if is_last else thn
        bg = fill_red if highlight else PatternFill(fill_type=None)
        ws.row_dimensions[row].height = 15

        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = font_normal
        lc.alignment = align_left
        lc.fill = bg
        lc.border = Border(left=med, right=thn, top=thn, bottom=bottom)
        for col in [2, 3]:
            ws.cell(row=row, column=col).border = Border(
                left=thn, right=thn, top=thn, bottom=bottom
            )
            ws.cell(row=row, column=col).fill = bg

        dc = ws.cell(row=row, column=4, value=b_val)
        dc.font = font_normal
        dc.alignment = align_right if isinstance(b_val, (int, float)) else align_center
        dc.fill = bg
        dc.border = Border(left=thn, right=thn, top=thn, bottom=bottom)

        ec = ws.cell(row=row, column=5, value=d_val)
        ec.font = font_normal
        ec.alignment = align_right if isinstance(d_val, (int, float)) else align_center
        ec.fill = bg
        ec.border = Border(left=thn, right=med, top=thn, bottom=bottom)

        return lc, dc, ec

    def write_list_col_headers(row, headers):
        ws.row_dimensions[row].height = 16
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == 5 else thn,
                top=thn,
                bottom=thn,
            )

    def write_list_data_rows(
        start_row,
        data_list,
        value_keys,
    ):

        num_rows = len(data_list)

        # =====================================
        # EMPTY TABLE SAFETY
        # =====================================
        if num_rows == 0:
            num_rows = 1
            data_list = [{}]

        for i in range(num_rows):
            r = start_row + i

            is_last = i == num_rows - 1

            bottom = med if is_last else thn

            ws.row_dimensions[r].height = 15

            row_data = data_list[i] if i < len(data_list) else {}

            vals = [row_data.get(k, "") for k in value_keys]

            for col in range(1, 6):
                val = vals[col - 1] if col - 1 < len(vals) else ""

                c = ws.cell(
                    row=r,
                    column=col,
                    value=val,
                )

                c.font = font_normal

                c.alignment = align_center

                c.fill = PatternFill(fill_type=None)

                c.border = Border(
                    left=med if col == 1 else thn,
                    right=med if col == 5 else thn,
                    top=thn,
                    bottom=bottom,
                )

        return start_row + num_rows

    # ── Layout ────────────────────────────────────────────────────────────────

    # Title — rows 1-2
    write_title(1)

    # Table 1: Files Reconciliation — rows 4-13
    write_section_header(4, "Files Reconciliation")
    write_col_headers_merged(5)

    recon_rows = [
        ("Total Cases Reported", False),
        ("Files Received", False),
        ("Files Pending", True),
        ("Files Incomplete", True),
        ("Files Verified", False),
        ("Files Approved", False),
        ("Files Rejected", True),
        ("Verification Completion %", False),
    ]
    for i, (item, highlight) in enumerate(recon_rows):
        r = 6 + i
        is_last = i == len(recon_rows) - 1
        b_val = booking.get(item, "")
        d_val = delivery.get(item, "")
        lc, dc, ec = write_data_row(r, item, b_val, d_val, highlight, is_last)
        if item == "Verification Completion %":
            dc.number_format = "0%"
            ec.number_format = "0%"

    # Table 2: Discount Summary — rows 16-26
    write_section_header(16, "Discount Summary")
    write_col_headers_merged(17)

    discount_rows = [
        ("Total Discount Given", False),
        ("Discount as per Approved Scheme", False),
        ("Net Excess Discount Amount", False),
        ("Highest Discount Car Model", False),
        ("Highest Discount Value", False),
        ("Excess Discount Cases", False),
        ("Allowable Discount Cases (out of Verified cases)", False),
        ("Excess Discount Cases(out of Verified cases)", True),
        ("Zero Discount Cases(out of Verified cases)", False),
    ]
    for i, (item, highlight) in enumerate(discount_rows):
        r = 18 + i
        is_last = i == len(discount_rows) - 1
        b_val = booking.get(item, "")
        d_val = delivery.get(item, "")
        lc, dc, ec = write_data_row(r, item, b_val, d_val, highlight, is_last)
        if any(
            kw in item
            for kw in ("Amount", "Value", "Discount Given", "Approved Scheme")
        ):
            dc.number_format = "##,##,##0"
            ec.number_format = "##,##,##0"

    # LIST 1 — PENDING FILES (BOOKING)
    booking_header_row = 28

    write_section_header(
        booking_header_row,
        "List of Pending Files (Booking)",
    )

    write_list_col_headers(
        booking_header_row + 1,
        [
            "S.No.",
            "Date",
            "Customer Name",
            "Mobile No.",
            "TL",
        ],
    )

    next_booking_pending_row = write_list_data_rows(
        start_row=booking_header_row + 2,
        data_list=booking_files_pending,
        value_keys=[
            "sno",
            "date",
            "name",
            "mobile",
            "tl",
        ],
    )

    # =========================================================
    # LIST 2 — OUT OF SCOPE (BOOKING)
    # =========================================================
    booking_oos_header_row = next_booking_pending_row + 2

    write_section_header(
        booking_oos_header_row,
        "Out Of Scope Files (Booking)",
    )

    write_list_col_headers(
        booking_oos_header_row + 1,
        [
            "S.No.",
            "Date",
            "Customer Name",
            "Mobile No.",
            "Reason",
        ],
    )

    next_booking_oos_row = write_list_data_rows(
        start_row=booking_oos_header_row + 2,
        data_list=booking_out_of_scope,
        value_keys=[
            "sno",
            "date",
            "name",
            "mobile",
            "reason",
        ],
    )
    # =========================================================
    # LIST 3 — DELAY IN RECEIVING FILES (BOOKING)
    # =========================================================
    booking_delay_header_row = next_booking_oos_row + 2

    write_section_header(
        booking_delay_header_row,
        "Delay In Receiving Files (Booking)",
    )

    write_list_col_headers(
        booking_delay_header_row + 1,
        [
            "S.No.",
            "Date",
            "Receiving Date",
            "Delay Days",
            "Customer Name",
        ],
    )

    next_booking_delay_row = write_list_data_rows(
        start_row=booking_delay_header_row + 2,
        data_list=booking_delay_files,
        value_keys=[
            "sno",
            "record_date",
            "receiving_date",
            "delay_days",
            "name",
        ],
    )

    # =========================================================
    # LIST 4 — PENDING FILES (DELIVERY)
    # =========================================================
    delivery_header_row = next_booking_delay_row + 2

    write_section_header(
        delivery_header_row,
        "List of Pending Files (Delivery)",
    )

    write_list_col_headers(
        delivery_header_row + 1,
        [
            "S.No.",
            "Date",
            "Customer Name",
            "Mobile No.",
            "TL",
        ],
    )

    next_delivery_pending_row = write_list_data_rows(
        start_row=delivery_header_row + 2,
        data_list=delivery_files_pending,
        value_keys=[
            "sno",
            "date",
            "name",
            "mobile",
            "tl",
        ],
    )

    # =========================================================
    # LIST 5 — OUT OF SCOPE (DELIVERY)
    # =========================================================
    delivery_oos_header_row = next_delivery_pending_row + 2

    write_section_header(
        delivery_oos_header_row,
        "Out Of Scope Files (Delivery)",
    )

    write_list_col_headers(
        delivery_oos_header_row + 1,
        [
            "S.No.",
            "Date",
            "Customer Name",
            "Mobile No.",
            "Reason",
        ],
    )

    next_delivery_oos_row = write_list_data_rows(
        start_row=delivery_oos_header_row + 2,
        data_list=delivery_out_of_scope,
        value_keys=[
            "sno",
            "date",
            "name",
            "mobile",
            "reason",
        ],
    )

    # =========================================================
    # LIST 6 — DELAY IN RECEIVING FILES (DELIVERY)
    # =========================================================
    delivery_delay_header_row = next_delivery_oos_row + 2

    write_section_header(
        delivery_delay_header_row,
        "Delay In Receiving Files (Delivery)",
    )

    write_list_col_headers(
        delivery_delay_header_row + 1,
        [
            "S.No.",
            "Date",
            "Receiving Date",
            "Delay Days",
            "Customer Name",
        ],
    )

    next_delivery_delay_row = write_list_data_rows(
        start_row=delivery_delay_header_row + 2,
        data_list=delivery_delay_files,
        value_keys=[
            "sno",
            "record_date",
            "receiving_date",
            "delay_days",
            "name",
        ],
    )
    # =========================================================
    # LIST 7 — REJECTED FILES DELIVERED
    # =========================================================
    rejected_delivery_header_row = next_delivery_delay_row + 2

    write_section_header(
        rejected_delivery_header_row,
        "Rejected Files Delivered",
    )

    write_list_col_headers(
        rejected_delivery_header_row + 1,
        [
            "S.No.",
            "Date",
            "Customer Name",
            "Mobile No.",
            "Reason",
        ],
    )

    write_list_data_rows(
        start_row=rejected_delivery_header_row + 2,
        data_list=rejected_files_delivered,
        value_keys=[
            "sno",
            "date",
            "name",
            "mobile",
            "reason",
        ],
    )

    # ═════════════════════════════════════════════════════════════════════════
    # ── Sheet 2: Pending Documents ────────────────────────────────────────────
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Pending Documents")

    # Checklist column definitions ──────────────────────────────────────────
    BOOKING_DOC_COLS = [
        "Customer KYC",
        "Vehicle Details",
        "Price Quotation",
        "Receipts",
        "Accessories Indent",
        "Exchange Details",
        "MD Reference Approval",
        "Corp ID",
        "Customer Sign",
    ]
    BOOKING_DOC_KEYS = [
        "kyc",
        "vehicle",
        "quotation",
        "receipts",
        "accessories_indent",
        "exchange",
        "md_approval",
        "corp_id",
        "customer_sign",
    ]

    DELIVERY_DOC_COLS = [
        "Customer Ledger",
        "Tax Invoice",
        "Accessories Indent",
        "Insurance",
        "RTO",
        "Finance",
        "Evaluation Certificate",
    ]
    DELIVERY_DOC_KEYS = [
        "ledger",
        "tax_invoice",
        "accessories_indent",
        "insurance",
        "rto",
        "finance",
        "eval_cert",
    ]

    BASE_COLS = ["S.No.", "Date", "Customer Name", "Mobile No.", "TL"]
    BASE_KEYS = ["sno", "date", "name", "mobile", "tl"]

    BOOKING_ALL_COLS = BASE_COLS + BOOKING_DOC_COLS  # 14 columns
    BOOKING_ALL_KEYS = BASE_KEYS + BOOKING_DOC_KEYS

    DELIVERY_ALL_COLS = BASE_COLS + DELIVERY_DOC_COLS  # 12 columns
    DELIVERY_ALL_KEYS = BASE_KEYS + DELIVERY_DOC_KEYS

    NUM_BOOKING_COLS = len(BOOKING_ALL_COLS)  # 14
    NUM_DELIVERY_COLS = len(DELIVERY_ALL_COLS)  # 12
    NUM_BASE_COLS = len(BASE_COLS)  # 5

    # Sheet 2 column widths ─────────────────────────────────────────────────
    ws2.column_dimensions["A"].width = 5  # S.no
    ws2.column_dimensions["B"].width = 11  # Date
    ws2.column_dimensions["C"].width = 20  # Customer Name
    ws2.column_dimensions["D"].width = 13  # Mobile No.
    ws2.column_dimensions["E"].width = 12  # TL
    # Checklist columns F onward (up to N = col 14)
    for ci in range(6, NUM_BOOKING_COLS + 1):
        ws2.column_dimensions[_col_letter(ci)].width = 14

    # Status colours — Excel-standard conditional-formatting palette ──────────
    # Each status: (background hex, font-colour hex)
    _STATUS_STYLES = {
        "received": ("C6EFCE", "006100"),  # light green  / dark green
        "pending": ("FFC7CE", "9C0006"),  # light red    / dark red
        "partial": ("FFEB9C", "9C6500"),  # light yellow / dark orange
        "na": ("BDD7EE", "1F4E79"),  # light blue   / dark blue
    }

    def _status_fill(val):
        entry = _STATUS_STYLES.get(str(val).strip().lower())
        if entry:
            return PatternFill("solid", fgColor=entry[0])
        return PatternFill(fill_type=None)

    def _status_font(val):
        entry = _STATUS_STYLES.get(str(val).strip().lower())
        if entry:
            return Font(name="Times New Roman", bold=True, color=entry[1])
        return font_normal

    # ── Sheet 2 helper functions ────────────────────────────────────────────

    def ws2_write_title(row_start):
        last = _col_letter(NUM_BOOKING_COLS)
        ws2.merge_cells(f"A{row_start}:{last}{row_start + 1}")
        c = ws2.cell(row=row_start, column=1, value=f"Daily Report As on ({date_str})")
        c.font = font_head
        c.alignment = align_center
        c.fill = fill_grey
        for r in range(row_start, row_start + 2):
            ws2.row_dimensions[r].height = 20
            for col in range(1, NUM_BOOKING_COLS + 1):
                ws2.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else Side(style=None),
                    right=med if col == NUM_BOOKING_COLS else Side(style=None),
                    top=med if r == row_start else Side(style=None),
                    bottom=med if r == row_start + 1 else Side(style=None),
                )

    def ws2_write_section_header(row, title, num_cols):
        last = _col_letter(num_cols)
        ws2.merge_cells(f"A{row}:{last}{row}")
        c = ws2.cell(row=row, column=1, value=title)
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_hdr
        ws2.row_dimensions[row].height = 18
        for col in range(1, num_cols + 1):
            ws2.cell(row=row, column=col).border = Border(
                left=med if col == 1 else Side(style=None),
                right=med if col == num_cols else Side(style=None),
                top=med,
                bottom=med,
            )

    def ws2_write_col_headers(row, headers):
        num_cols = len(headers)
        ws2.row_dimensions[row].height = 30
        for col, val in enumerate(headers, start=1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == num_cols else thn,
                top=thn,
                bottom=thn,
            )

    def ws2_write_data_rows(
        start_row,
        data_list,
        value_keys,
        num_cols,
    ):

        num_rows = len(data_list)

        # =====================================
        # EMPTY TABLE SAFETY
        # =====================================
        if num_rows == 0:
            num_rows = 1
            data_list = [{}]

        for i in range(num_rows):
            r = start_row + i

            is_last = i == num_rows - 1

            bottom = med if is_last else thn

            ws2.row_dimensions[r].height = 15

            row_data = data_list[i] if i < len(data_list) else {}

            vals = [row_data.get(k, "") for k in value_keys]

            for col in range(1, num_cols + 1):
                val = vals[col - 1] if col - 1 < len(vals) else ""

                c = ws2.cell(
                    row=r,
                    column=col,
                    value=val,
                )

                if col > NUM_BASE_COLS:
                    c.fill = _status_fill(val)

                    c.font = _status_font(val)

                else:
                    c.fill = PatternFill(fill_type=None)

                    c.font = font_normal

                c.alignment = align_center

                c.border = Border(
                    left=med if col == 1 else thn,
                    right=med if col == num_cols else thn,
                    top=thn,
                    bottom=bottom,
                )

        return start_row + num_rows

    # ── Sheet 2 Layout ──────────────────────────────────────────────────────

    # Title — rows 1-2
    ws2_write_title(1)

    # BOOKING DOCS PENDING
    booking_docs_header_row = 4

    ws2_write_section_header(
        booking_docs_header_row,
        "List of Pending Documents (Booking)",
        NUM_BOOKING_COLS,
    )

    ws2_write_col_headers(
        booking_docs_header_row + 1,
        BOOKING_ALL_COLS,
    )

    next_ws2_row = ws2_write_data_rows(
        start_row=booking_docs_header_row + 2,
        data_list=booking_docs_pending,
        value_keys=BOOKING_ALL_KEYS,
        num_cols=NUM_BOOKING_COLS,
    )

    # DELIVERY DOCS PENDING
    delivery_docs_header_row = next_ws2_row + 2

    ws2_write_section_header(
        delivery_docs_header_row,
        "List of Pending Documents (Delivery)",
        NUM_DELIVERY_COLS,
    )

    ws2_write_col_headers(
        delivery_docs_header_row + 1,
        DELIVERY_ALL_COLS,
    )

    ws2_write_data_rows(
        start_row=delivery_docs_header_row + 2,
        data_list=delivery_docs_pending,
        value_keys=DELIVERY_ALL_KEYS,
        num_cols=NUM_DELIVERY_COLS,
    )

    # ── Save ──────────────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_date = date_str.replace("/", "-") if date_str else "no-date"
    file_name = f"Daily-Report-{safe_date}.xlsx"
    return buffer, file_name


# ── Utility ───────────────────────────────────────────────────────────────────
def _col_letter(n):
    """Convert 1-based column number to Excel letter(s). e.g. 1→A, 14→N, 27→AA."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result
