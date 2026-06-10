"""
report_generator.py
════════════════════════════════════════════════════════════════════════════════
Excel Report Generator — Daily & Combined (N-Dealer)
────────────────────────────────────────────────────────────────────────────────
PUBLIC API
──────────
    generate_report(report_type, backend_data)  ← single entry point for frontend

    report_type values:
        "daily"    → one dealer, 5-column layout  (Sheet 1 + Sheet 2 docs)
        "combine"  → N dealers, dynamic columns   (single sheet, N×2 data cols)

    Returns: (BytesIO, filename_str)

FRONTEND INTEGRATION EXAMPLE (Flask)
──────────────────────────────────────
    from flask import Flask, request, send_file
    from report_generator import generate_report

    app = Flask(__name__)

    @app.route("/api/report", methods=["POST"])
    def download_report():
        payload      = request.get_json()
        report_type  = payload.get("report_type")   # "daily" | "combine"
        backend_data = payload.get("data", {})
        buffer, filename = generate_report(report_type, backend_data)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

────────────────────────────────────────────────────────────────────────────────
BACKEND_DATA SCHEMAS
════════════════════════════════════════════════════════════════════════════════

── "daily" report ─────────────────────────────────────────────────────────────

    {
        "report_date": "30/04/2026",        # str dd/mm/yyyy | date | datetime
                                            # or {"from": "01/04/2026", "to": "30/04/2026"}
        "scope": "Booking",                 # label shown in title & sheet tab

        "booking":  { <METRIC_KEYS> },
        "delivery": { <METRIC_KEYS> },

        # ── Sheet 1 lists (all optional; empty list → one blank row) ──────────
        "booking_files_pending":    [ {sno, date, name, mobile, tl}, ... ],
        "delivery_files_pending":   [ {sno, date, name, mobile, tl}, ... ],
        "booking_out_of_scope":     [ {sno, date, name, mobile, reason}, ... ],
        "delivery_out_of_scope":    [ {sno, date, name, mobile, reason}, ... ],
        "booking_delay_files":      [ {sno, record_date, receiving_date, delay_days, name}, ... ],
        "delivery_delay_files":     [ {sno, record_date, receiving_date, delay_days, name}, ... ],
        "rejected_files_delivered": [ {sno, date, name, mobile, reason}, ... ],

        # ── Sheet 2 doc-checklist lists ────────────────────────────────────
        "booking_docs_pending": [
            {
                sno, date, name, mobile, tl,
                kyc, vehicle, quotation, receipts,
                accessories_indent, exchange, md_approval,
                corp_id, customer_sign                      # values: Received|Pending|Partial|NA
            }, ...
        ],
        "delivery_docs_pending": [
            {
                sno, date, name, mobile, tl,
                ledger, tax_invoice, accessories_indent,
                insurance, rto, finance, eval_cert          # values: Received|Pending|Partial|NA
            }, ...
        ],
    }

── "combine" report ────────────────────────────────────────────────────────────

    {
        "report_date": "30/04/2026",        # same formats as daily
        "scope": "All Dealers",             # label shown in title

        "dealers": [                        # 1 or more dealers — columns auto-expand
            {
                "name":     "Dealer Name 1",
                "booking":  { <METRIC_KEYS> },
                "delivery": { <METRIC_KEYS> },
            },
            {
                "name":     "Dealer Name 2",
                "booking":  { <METRIC_KEYS> },
                "delivery": { <METRIC_KEYS> },
            },
            # ...add as many dealers as needed
        ],
    }

── METRIC_KEYS (used by both report types) ─────────────────────────────────────

    # Files Reconciliation table
    "Total Cases Reported"                              : int
    "Files Received"                                    : int
    "Files Pending"                                     : int          ← red highlight
    "Files Incomplete"                                  : int          ← red highlight
    "Files Verified"                                    : int
    "Files Approved"                                    : int
    "Files Rejected"                                    : int          ← red highlight
    "Verification Completion %"                         : float 0-1    ← shown as 0%..100%

    # Discount Summary table
    "Total Discount Given"                              : int/float    ← Indian comma format
    "Discount as per Approved Scheme"                   : int/float    ← Indian comma format
    "Net Excess Discount Amount"                        : int/float    ← Indian comma format
    "Highest Discount Car Model"                        : str
    "Highest Discount Value"                            : int/float    ← Indian comma format
    "Excess Discount Cases"                             : int
    "Allowable Discount Cases (out of Verified cases)"  : int
    "Excess Discount Cases(out of Verified cases)"      : int          ← red highlight
    "Zero Discount Cases(out of Verified cases)"        : int

════════════════════════════════════════════════════════════════════════════════
"""

from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# PUBLIC ENTRY POINT
def generate_report(report_type: str, backend_data: dict) -> tuple:
    """
    Unified dispatcher — the only function the frontend/backend route needs to call.

    Parameters
    ----------
    report_type : "daily" | "combine"
    backend_data : dict — see module docstring for full schema

    Returns
    -------
    (BytesIO, str)  — file buffer ready for send_file / streaming + filename
    """
    report_type = (report_type or "").strip().lower()

    if report_type == "daily":
        return _generate_daily_report(backend_data)
    elif report_type == "combine":
        return _generate_combine_report(backend_data)
    else:
        raise ValueError(
            f"Unknown report_type '{report_type}'. Expected 'daily' or 'combine'."
        )


# SHARED INTERNALS
def _col_letter(n: int) -> str:
    """1-based column index → Excel column letter. e.g. 1→A, 27→AA."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _resolve_date_str(raw) -> str:
    """Normalise any date input to a display string."""
    if isinstance(raw, (date, datetime)):
        return raw.strftime("%d/%m/%Y")
    if isinstance(raw, dict):
        frm = raw.get("from")
        to = raw.get("to")
        if isinstance(frm, (date, datetime)):
            frm = frm.strftime("%d/%m/%Y")
        if isinstance(to, (date, datetime)):
            to = to.strftime("%d/%m/%Y")
        if frm and to:
            return f"{frm} To {to}"
        return str(frm or to or "")
    return str(raw) if raw else ""


def _styles() -> dict:
    """Return a shared set of openpyxl style objects."""
    return dict(
        font_bold=Font(name="Times New Roman", bold=True, color="000000"),
        font_head=Font(name="Times New Roman", bold=True, size=15, color="000000"),
        font_normal=Font(name="Times New Roman", bold=False, color="000000"),
        align_center=Alignment(horizontal="center", vertical="center", wrap_text=True),
        align_left=Alignment(horizontal="left", vertical="center", wrap_text=True),
        align_right=Alignment(horizontal="right", vertical="center"),
        med=Side(style="medium"),
        thn=Side(style="thin"),
        fill_grey=PatternFill("solid", fgColor="E7E6E6"),
        fill_blue_hdr=PatternFill("solid", fgColor="9BC2E6"),
        fill_blue_sub=PatternFill("solid", fgColor="D9E1F2"),
        fill_red=PatternFill("solid", fgColor="FF6D6D"),
        fill_none=PatternFill(fill_type=None),
    )


# Rows that are red-highlighted in the Files Reconciliation table
_RECON_HIGHLIGHT = {"Files Pending", "Files Incomplete", "Files Rejected"}

# Rows that are red-highlighted in the Discount Summary table
_DISCOUNT_HIGHLIGHT = {"Excess Discount Cases(out of Verified cases)"}

# Ordered row definitions for both tables
_RECON_ROWS = [
    "Total Cases Reported",
    "Files Received",
    "Files Pending",
    "Files Incomplete",
    "Files Verified",
    "Files Approved",
    "Files Rejected",
    "Verification Completion %",
]

_DISCOUNT_ROWS = [
    "Total Discount Given",
    "Discount as per Approved Scheme",
    "Net Excess Discount Amount",
    "Highest Discount Car Model",
    "Highest Discount Value",
    "Excess Discount Cases",
    "Allowable Discount Cases (out of Verified cases)",
    "Excess Discount Cases(out of Verified cases)",
    "Zero Discount Cases(out of Verified cases)",
]

# Keys whose values get Indian comma number format
_MONEY_KEYWORDS = ("Amount", "Value", "Discount Given", "Approved Scheme")


# REPORT 1 — DAILY  (5 columns, 2 sheets)


def _generate_daily_report(backend_data: dict) -> tuple:
    if backend_data is None:
        backend_data = {}

    report_scope = backend_data.get("scope", "")
    date_str = _resolve_date_str(backend_data.get("report_date", ""))
    booking = backend_data.get("booking", {})
    delivery = backend_data.get("delivery", {})

    # Sheet 1 lists
    booking_files_pending = backend_data.get("booking_files_pending", [])
    delivery_files_pending = backend_data.get("delivery_files_pending", [])
    booking_out_of_scope = backend_data.get("booking_out_of_scope", [])
    delivery_out_of_scope = backend_data.get("delivery_out_of_scope", [])
    booking_delay_files = backend_data.get("booking_delay_files", [])
    delivery_delay_files = backend_data.get("delivery_delay_files", [])
    rejected_files_delivered = backend_data.get("rejected_files_delivered", [])
    # Sheet 2 lists
    booking_docs_pending = backend_data.get("booking_docs_pending", [])
    delivery_docs_pending = backend_data.get("delivery_docs_pending", [])

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Report for {report_scope}"

    S = _styles()
    font_bold = S["font_bold"]
    font_head = S["font_head"]
    font_normal = S["font_normal"]
    align_center = S["align_center"]
    align_left = S["align_left"]
    align_right = S["align_right"]
    med = S["med"]
    thn = S["thn"]
    fill_grey = S["fill_grey"]
    fill_blue_hdr = S["fill_blue_hdr"]
    fill_blue_sub = S["fill_blue_sub"]
    fill_red = S["fill_red"]

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 33
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # ── Sheet-1 helper functions ─────────────────────────────────────────────
    def s1_title(row_start):
        ws.merge_cells(f"A{row_start}:E{row_start + 1}")
        c = ws.cell(
            row=row_start,
            column=1,
            value=f"Report for {report_scope}\nAs on ({date_str})",
        )
        c.font = font_head
        c.alignment = align_center
        c.fill = fill_grey
        for r in range(row_start, row_start + 2):
            ws.row_dimensions[r].height = 20
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else Side(style=None),
                    right=med if col == 5 else Side(style=None),
                    top=med if r == row_start else Side(style=None),
                    bottom=med if r == row_start + 1 else Side(style=None),
                )

    def s1_section_hdr(row, title):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_hdr
        ws.row_dimensions[row].height = 18
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(
                left=med if col == 1 else Side(style=None),
                right=med if col == 5 else Side(style=None),
                top=med,
                bottom=med,
            )

    def s1_col_headers(row, d_label="Booking", e_label="Delivery"):
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 16
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == 5 else thn,
                top=thn,
                bottom=thn,
            )
        ws.cell(row=row, column=1).value = "Particulars"
        ws.cell(row=row, column=4).value = d_label
        ws.cell(row=row, column=5).value = e_label

    def s1_data_row(row, label, b_val, d_val, highlight=False, is_last=False):
        bottom = med if is_last else thn
        bg = fill_red if highlight else PatternFill(fill_type=None)
        ws.row_dimensions[row].height = 15
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = font_normal
        lc.alignment = align_left
        lc.fill = bg
        lc.border = Border(left=med, right=thn, top=thn, bottom=bottom)
        for col in [2, 3]:
            ws.cell(row=row, column=col).fill = bg
            ws.cell(row=row, column=col).border = Border(
                left=thn, right=thn, top=thn, bottom=bottom
            )
        dc = ws.cell(row=row, column=4, value=b_val)
        dc.font = font_normal
        dc.alignment = align_right if isinstance(b_val, (int, float)) else align_center
        dc.fill = bg
        dc.border = Border(left=thn, right=thn, top=thn, bottom=bottom)
        ec = ws.cell(row=row, column=5, value=d_val)
        ec.font = font_normal
        ec.alignment = align_right if isinstance(d_val, (int, float)) else align_center
        ec.fill = bg
        ec.border = Border(left=thn, right=med, top=thn, bottom=bottom)
        return lc, dc, ec

    def s1_list_col_headers(row, headers):
        ws.row_dimensions[row].height = 16
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == 5 else thn,
                top=thn,
                bottom=thn,
            )

    def s1_list_data_rows(start_row, data_list, keys):
        rows = data_list if data_list else [{}]
        for i, row_data in enumerate(rows):
            r = start_row + i
            is_last = i == len(rows) - 1
            bottom = med if is_last else thn
            ws.row_dimensions[r].height = 15
            for col, key in enumerate(keys, start=1):
                c = ws.cell(row=r, column=col, value=row_data.get(key, ""))
                c.font = font_normal
                c.alignment = align_center
                c.fill = PatternFill(fill_type=None)
                c.border = Border(
                    left=med if col == 1 else thn,
                    right=med if col == 5 else thn,
                    top=thn,
                    bottom=bottom,
                )
        return start_row + len(rows)

    # ── Sheet 1 Layout ───────────────────────────────────────────────────────
    s1_title(1)

    # Files Reconciliation
    s1_section_hdr(4, "Files Reconciliation")
    s1_col_headers(5)
    for i, item in enumerate(_RECON_ROWS):
        r = 6 + i
        hl = item in _RECON_HIGHLIGHT
        is_last = i == len(_RECON_ROWS) - 1
        lc, dc, ec = s1_data_row(
            r, item, booking.get(item, ""), delivery.get(item, ""), hl, is_last
        )
        if item == "Verification Completion %":
            dc.number_format = "0%"
            ec.number_format = "0%"

    # Discount Summary
    s1_section_hdr(16, "Discount Summary")
    s1_col_headers(17)
    for i, item in enumerate(_DISCOUNT_ROWS):
        r = 18 + i
        hl = item in _DISCOUNT_HIGHLIGHT
        is_last = i == len(_DISCOUNT_ROWS) - 1
        lc, dc, ec = s1_data_row(
            r, item, booking.get(item, ""), delivery.get(item, ""), hl, is_last
        )
        if any(kw in item for kw in _MONEY_KEYWORDS):
            dc.number_format = "##,##,##0"
            ec.number_format = "##,##,##0"

    # ── Sheet 1 List sections ────────────────────────────────────────────────
    cur = 28

    sections = [
        (
            "List of Pending Files (Booking)",
            ["S.No.", "Date", "Customer Name", "Mobile No.", "TL"],
            ["sno", "date", "name", "mobile", "tl"],
            booking_files_pending,
        ),
        (
            "Out Of Scope Files (Booking)",
            ["S.No.", "Date", "Customer Name", "Mobile No.", "Reason"],
            ["sno", "date", "name", "mobile", "reason"],
            booking_out_of_scope,
        ),
        (
            "Delay In Receiving Files (Booking)",
            ["S.No.", "Date", "Receiving Date", "Delay Days", "Customer Name"],
            ["sno", "record_date", "receiving_date", "delay_days", "name"],
            booking_delay_files,
        ),
        (
            "List of Pending Files (Delivery)",
            ["S.No.", "Date", "Customer Name", "Mobile No.", "TL"],
            ["sno", "date", "name", "mobile", "tl"],
            delivery_files_pending,
        ),
        (
            "Out Of Scope Files (Delivery)",
            ["S.No.", "Date", "Customer Name", "Mobile No.", "Reason"],
            ["sno", "date", "name", "mobile", "reason"],
            delivery_out_of_scope,
        ),
        (
            "Delay In Receiving Files (Delivery)",
            ["S.No.", "Date", "Receiving Date", "Delay Days", "Customer Name"],
            ["sno", "record_date", "receiving_date", "delay_days", "name"],
            delivery_delay_files,
        ),
        (
            "Rejected Files Delivered",
            ["S.No.", "Date", "Customer Name", "Mobile No.", "Reason"],
            ["sno", "date", "name", "mobile", "reason"],
            rejected_files_delivered,
        ),
    ]

    for title, headers, keys, data in sections:
        s1_section_hdr(cur, title)
        s1_list_col_headers(cur + 1, headers)
        cur = s1_list_data_rows(cur + 2, data, keys) + 2

    # ── Sheet 2: Pending Documents ───────────────────────────────────────────
    ws2 = wb.create_sheet(title="Pending Documents")

    BOOKING_DOC_COLS = [
        "Customer KYC",
        "Vehicle Details",
        "Price Quotation",
        "Receipts",
        "Accessories Indent",
        "Exchange Details",
        "MD Reference Approval",
        "Corp ID",
        "Customer Sign",
    ]
    BOOKING_DOC_KEYS = [
        "kyc",
        "vehicle",
        "quotation",
        "receipts",
        "accessories_indent",
        "exchange",
        "md_approval",
        "corp_id",
        "customer_sign",
    ]
    DELIVERY_DOC_COLS = [
        "Customer Ledger",
        "Tax Invoice",
        "Accessories Indent",
        "Insurance",
        "RTO",
        "Finance",
        "Evaluation Certificate",
    ]
    DELIVERY_DOC_KEYS = [
        "ledger",
        "tax_invoice",
        "accessories_indent",
        "insurance",
        "rto",
        "finance",
        "eval_cert",
    ]
    BASE_COLS = ["S.No.", "Date", "Customer Name", "Mobile No.", "TL"]
    BASE_KEYS = ["sno", "date", "name", "mobile", "tl"]
    B_COLS = BASE_COLS + BOOKING_DOC_COLS
    B_KEYS = BASE_KEYS + BOOKING_DOC_KEYS
    D_COLS = BASE_COLS + DELIVERY_DOC_COLS
    D_KEYS = BASE_KEYS + DELIVERY_DOC_KEYS
    NB = len(B_COLS)
    ND = len(D_COLS)
    NBase = len(BASE_COLS)

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 11
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 13
    ws2.column_dimensions["E"].width = 12
    for ci in range(6, NB + 1):
        ws2.column_dimensions[_col_letter(ci)].width = 14

    STATUS_STYLES = {
        "received": ("C6EFCE", "006100"),
        "pending": ("FFC7CE", "9C0006"),
        "partial": ("FFEB9C", "9C6500"),
        "na": ("BDD7EE", "1F4E79"),
    }

    def _sfill(val):
        e = STATUS_STYLES.get(str(val).strip().lower())
        return PatternFill("solid", fgColor=e[0]) if e else PatternFill(fill_type=None)

    def _sfont(val):
        e = STATUS_STYLES.get(str(val).strip().lower())
        return Font(name="Times New Roman", bold=True, color=e[1]) if e else font_normal

    def s2_title(row_start):
        last = _col_letter(NB)
        ws2.merge_cells(f"A{row_start}:{last}{row_start + 1}")
        c = ws2.cell(row=row_start, column=1, value=f"Daily Report As on ({date_str})")
        c.font = font_head
        c.alignment = align_center
        c.fill = fill_grey
        for r in range(row_start, row_start + 2):
            ws2.row_dimensions[r].height = 20
            for col in range(1, NB + 1):
                ws2.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else Side(style=None),
                    right=med if col == NB else Side(style=None),
                    top=med if r == row_start else Side(style=None),
                    bottom=med if r == row_start + 1 else Side(style=None),
                )

    def s2_section_hdr(row, title, num_cols):
        last = _col_letter(num_cols)
        ws2.merge_cells(f"A{row}:{last}{row}")
        c = ws2.cell(row=row, column=1, value=title)
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_hdr
        ws2.row_dimensions[row].height = 18
        for col in range(1, num_cols + 1):
            ws2.cell(row=row, column=col).border = Border(
                left=med if col == 1 else Side(style=None),
                right=med if col == num_cols else Side(style=None),
                top=med,
                bottom=med,
            )

    def s2_col_headers(row, headers):
        nc = len(headers)
        ws2.row_dimensions[row].height = 30
        for col, val in enumerate(headers, start=1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            c.border = Border(
                left=med if col == 1 else thn,
                right=med if col == nc else thn,
                top=thn,
                bottom=thn,
            )

    def s2_data_rows(start_row, data_list, keys, num_cols):
        rows = data_list if data_list else [{}]
        for i, row_data in enumerate(rows):
            r = start_row + i
            is_last = i == len(rows) - 1
            bottom = med if is_last else thn
            ws2.row_dimensions[r].height = 15
            for col, key in enumerate(keys, start=1):
                val = row_data.get(key, "")
                c = ws2.cell(row=r, column=col, value=val)
                c.fill = _sfill(val) if col > NBase else PatternFill(fill_type=None)
                c.font = _sfont(val) if col > NBase else font_normal
                c.alignment = align_center
                c.border = Border(
                    left=med if col == 1 else thn,
                    right=med if col == num_cols else thn,
                    top=thn,
                    bottom=bottom,
                )
        return start_row + len(rows)

    s2_title(1)
    s2_section_hdr(4, "List of Pending Documents (Booking)", NB)
    s2_col_headers(5, B_COLS)
    next_row = s2_data_rows(6, booking_docs_pending, B_KEYS, NB)

    ddr = next_row + 2
    s2_section_hdr(ddr, "List of Pending Documents (Delivery)", ND)
    s2_col_headers(ddr + 1, D_COLS)
    s2_data_rows(ddr + 2, delivery_docs_pending, D_KEYS, ND)

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = date_str.replace("/", "-") if date_str else "no-date"
    return buf, f"Daily-Report-{safe}.xlsx"


# REPORT 2 — COMBINE  (N dealers, dynamic columns, 1 sheet)


def _generate_combine_report(backend_data: dict) -> tuple:
    if backend_data is None:
        backend_data = {}

    report_scope = backend_data.get("scope", "All Dealers")
    date_str = _resolve_date_str(backend_data.get("report_date", ""))
    dealers = backend_data.get("dealers", [])

    # Guarantee at least 1 dealer so the layout doesn't break
    if not dealers:
        dealers = [{"name": "Dealer 1", "booking": {}, "delivery": {}}]

    num_dealers = len(dealers)
    # Column layout:
    #   A B C  — label (always 3 cols, merged)
    #   D E    — Dealer 1 Booking | Delivery
    #   F G    — Dealer 2 Booking | Delivery
    #   ...    — Dealer N Booking | Delivery
    NUM_COLS = 3 + num_dealers * 2
    LAST_COL = _col_letter(NUM_COLS)

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Report"

    S = _styles()
    font_bold = S["font_bold"]
    font_head = S["font_head"]
    font_normal = S["font_normal"]
    align_center = S["align_center"]
    align_left = S["align_left"]
    align_right = S["align_right"]
    med = S["med"]
    thn = S["thn"]
    fill_grey = S["fill_grey"]
    fill_blue_hdr = S["fill_blue_hdr"]
    fill_blue_sub = S["fill_blue_sub"]
    fill_red = S["fill_red"]

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 33
    for di in range(num_dealers * 2):
        ws.column_dimensions[_col_letter(4 + di)].width = 18

    # ── Helper: standard cell border ────────────────────────────────────────
    def _bdr(col, top=thn, bottom=thn):
        return Border(
            left=med if col == 1 else thn,
            right=med if col == NUM_COLS else thn,
            top=top,
            bottom=bottom,
        )

    # ── Title (rows row_start and row_start+1) ───────────────────────────────
    def c_title(row_start):
        ws.merge_cells(f"A{row_start}:{LAST_COL}{row_start + 1}")
        c = ws.cell(
            row=row_start,
            column=1,
            value=f"Report for {report_scope}\nAs on ({date_str})",
        )
        c.font = font_head
        c.alignment = align_center
        c.fill = fill_grey
        for r in range(row_start, row_start + 2):
            ws.row_dimensions[r].height = 20
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else Side(style=None),
                    right=med if col == NUM_COLS else Side(style=None),
                    top=med if r == row_start else Side(style=None),
                    bottom=med if r == row_start + 1 else Side(style=None),
                )

    # ── Full-width section header ────────────────────────────────────────────
    def c_section_hdr(row, title):
        ws.merge_cells(f"A{row}:{LAST_COL}{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_hdr
        ws.row_dimensions[row].height = 18
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=row, column=col).border = Border(
                left=med if col == 1 else Side(style=None),
                right=med if col == NUM_COLS else Side(style=None),
                top=med,
                bottom=med,
            )

    # ── Two-row column header block ──────────────────────────────────────────
    # Row hdr   : [Particulars (A:C, rowspan 2)] [Dealer 1 (D:E)] [Dealer 2 (F:G)] ...
    # Row hdr+1 :                                [Booking][Delivery] per dealer
    def c_col_headers(hdr_row):
        # "Particulars" spanning 2 rows
        ws.merge_cells(f"A{hdr_row}:C{hdr_row + 1}")
        c = ws.cell(row=hdr_row, column=1, value="Particulars")
        c.font = font_bold
        c.alignment = align_center
        c.fill = fill_blue_sub
        for r in range(hdr_row, hdr_row + 2):
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = fill_blue_sub
                ws.cell(row=r, column=col).border = Border(
                    left=med if col == 1 else thn,
                    right=thn,
                    top=med if r == hdr_row else thn,
                    bottom=thn,
                )

        # Per-dealer pair
        for di, dealer in enumerate(dealers):
            b_col = 4 + di * 2  # Booking column index
            d_col = b_col + 1  # Delivery column index
            b_ltr = _col_letter(b_col)
            d_ltr = _col_letter(d_col)

            # Dealer name (merged across booking+delivery in row hdr)
            ws.merge_cells(f"{b_ltr}{hdr_row}:{d_ltr}{hdr_row}")
            c = ws.cell(
                row=hdr_row, column=b_col, value=dealer.get("name", f"Dealer {di + 1}")
            )
            c.font = font_bold
            c.alignment = align_center
            c.fill = fill_blue_sub
            for col in [b_col, d_col]:
                ws.cell(row=hdr_row, column=col).fill = fill_blue_sub
                ws.cell(row=hdr_row, column=col).border = Border(
                    left=thn,
                    right=med if col == NUM_COLS else thn,
                    top=med,
                    bottom=thn,
                )

            # Booking / Delivery sub-headers (row hdr+1)
            for col, label in [(b_col, "Booking"), (d_col, "Delivery")]:
                c = ws.cell(row=hdr_row + 1, column=col, value=label)
                c.font = font_bold
                c.alignment = align_center
                c.fill = fill_blue_sub
                c.border = Border(
                    left=thn,
                    right=med if col == NUM_COLS else thn,
                    top=thn,
                    bottom=thn,
                )

        ws.row_dimensions[hdr_row].height = 18
        ws.row_dimensions[hdr_row + 1].height = 16

    # ── Data row ─────────────────────────────────────────────────────────────
    def c_data_row(row, label, values, highlight=False, is_last=False):
        """
        values = flat list alternating booking/delivery per dealer:
                 [d1_b, d1_d, d2_b, d2_d, ...]
        """
        bottom = med if is_last else thn
        bg = fill_red if highlight else PatternFill(fill_type=None)
        ws.row_dimensions[row].height = 15

        # Label (A:C merged)
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = font_normal
        lc.alignment = align_left
        lc.fill = bg
        lc.border = Border(left=med, right=thn, top=thn, bottom=bottom)
        for col in [2, 3]:
            ws.cell(row=row, column=col).fill = bg
            ws.cell(row=row, column=col).border = Border(
                left=thn, right=thn, top=thn, bottom=bottom
            )

        # Value cells
        data_cells = []
        for vi, val in enumerate(values):
            col = 4 + vi
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_normal
            c.alignment = align_right if isinstance(val, (int, float)) else align_center
            c.fill = bg
            c.border = Border(
                left=thn,
                right=med if col == NUM_COLS else thn,
                top=thn,
                bottom=bottom,
            )
            data_cells.append(c)
        return lc, data_cells

    # ── Layout ───────────────────────────────────────────────────────────────
    c_title(1)  # rows 1-2

    # Files Reconciliation — header row 4, col-headers rows 5-6, data rows 7-14
    c_section_hdr(4, "Files Reconciliation")
    c_col_headers(5)
    for i, item in enumerate(_RECON_ROWS):
        r = 7 + i
        hl = item in _RECON_HIGHLIGHT
        is_last = i == len(_RECON_ROWS) - 1
        vals = []
        for d in dealers:
            vals.append(d.get("booking", {}).get(item, ""))
            vals.append(d.get("delivery", {}).get(item, ""))
        lc, dcs = c_data_row(r, item, vals, hl, is_last)
        if item == "Verification Completion %":
            for dc in dcs:
                dc.number_format = "0%"

    # Discount Summary — header row 16, col-headers rows 17-18, data rows 19-27
    c_section_hdr(16, "Discount Summary")
    c_col_headers(17)
    for i, item in enumerate(_DISCOUNT_ROWS):
        r = 19 + i
        hl = item in _DISCOUNT_HIGHLIGHT
        is_last = i == len(_DISCOUNT_ROWS) - 1
        vals = []
        for d in dealers:
            vals.append(d.get("booking", {}).get(item, ""))
            vals.append(d.get("delivery", {}).get(item, ""))
        lc, dcs = c_data_row(r, item, vals, hl, is_last)
        if any(kw in item for kw in _MONEY_KEYWORDS):
            for dc in dcs:
                dc.number_format = "##,##,##0"

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = date_str.replace("/", "-") if date_str else "no-date"
    return buf, f"Combined-Report-{safe}.xlsx"


# SAMPLE DATA + SELF-TEST  (run: python report_generator.py)

if __name__ == "__main__":
    import os

    METRIC = {
        "Total Cases Reported": 100,
        "Files Received": 95,
        "Files Pending": 7,
        "Files Incomplete": 2,
        "Files Verified": 86,
        "Files Approved": 80,
        "Files Rejected": 6,
        "Verification Completion %": 0.86,
        "Total Discount Given": 450000,
        "Discount as per Approved Scheme": 400000,
        "Net Excess Discount Amount": 50000,
        "Highest Discount Car Model": "Swift Dzire",
        "Highest Discount Value": 30000,
        "Excess Discount Cases": 10,
        "Allowable Discount Cases (out of Verified cases)": 76,
        "Excess Discount Cases(out of Verified cases)": 10,
        "Zero Discount Cases(out of Verified cases)": 3,
    }

    def scaled(n):
        return {
            k: (
                round(v * n, 2)
                if isinstance(v, float)
                else (v + " " + str(n) if isinstance(v, str) else int(v * n))
            )
            for k, v in METRIC.items()
        }

    # ── daily ──────────────────────────────────────────────────────────────
    daily_data = {
        "report_date": "22/05/2026",
        "scope": "Booking",
        "booking": scaled(1),
        "delivery": scaled(0.8),
        "booking_files_pending": [
            {
                "sno": 1,
                "date": "20/05/2026",
                "name": "Ramesh Kumar",
                "mobile": "9876543210",
                "tl": "Amit S",
            },
            {
                "sno": 2,
                "date": "21/05/2026",
                "name": "Sunita Devi",
                "mobile": "9123456780",
                "tl": "Priya M",
            },
        ],
        "delivery_files_pending": [
            {
                "sno": 1,
                "date": "19/05/2026",
                "name": "Anil Sharma",
                "mobile": "9988776655",
                "tl": "Rohit K",
            },
        ],
        "booking_out_of_scope": [],
        "delivery_out_of_scope": [],
        "booking_delay_files": [],
        "delivery_delay_files": [],
        "rejected_files_delivered": [],
        "booking_docs_pending": [
            {
                "sno": 1,
                "date": "20/05/2026",
                "name": "Ramesh Kumar",
                "mobile": "9876543210",
                "tl": "Amit S",
                "kyc": "Received",
                "vehicle": "Pending",
                "quotation": "Received",
                "receipts": "Partial",
                "accessories_indent": "NA",
                "exchange": "Pending",
                "md_approval": "Received",
                "corp_id": "NA",
                "customer_sign": "Received",
            },
        ],
        "delivery_docs_pending": [
            {
                "sno": 1,
                "date": "19/05/2026",
                "name": "Anil Sharma",
                "mobile": "9988776655",
                "tl": "Rohit K",
                "ledger": "Received",
                "tax_invoice": "Pending",
                "accessories_indent": "NA",
                "insurance": "Received",
                "rto": "Partial",
                "finance": "Received",
                "eval_cert": "NA",
            },
        ],
    }

    # ── combine — 1 dealer ─────────────────────────────────────────────────
    combine_1 = {
        "report_date": "22/05/2026",
        "scope": "Single Dealer",
        "dealers": [
            {"name": "Dealer Alpha", "booking": scaled(1), "delivery": scaled(0.8)}
        ],
    }

    # ── combine — 2 dealers ────────────────────────────────────────────────
    combine_2 = {
        "report_date": "22/05/2026",
        "scope": "All Dealers",
        "dealers": [
            {"name": "Dealer Alpha", "booking": scaled(1), "delivery": scaled(0.8)},
            {"name": "Dealer Beta", "booking": scaled(0.7), "delivery": scaled(0.6)},
        ],
    }

    # ── combine — 3 dealers ────────────────────────────────────────────────
    combine_3 = {
        "report_date": "22/05/2026",
        "scope": "All Dealers",
        "dealers": [
            {"name": "Dealer Alpha", "booking": scaled(1), "delivery": scaled(0.8)},
            {"name": "Dealer Beta", "booking": scaled(0.7), "delivery": scaled(0.6)},
            {"name": "Dealer Gamma", "booking": scaled(0.5), "delivery": scaled(0.4)},
        ],
    }

    OUT = "/home/claude"
    tests = [
        ("daily", daily_data),
        ("combine", combine_1),
        ("combine", combine_2),
        ("combine", combine_3),
    ]
    for rtype, data in tests:
        buf, fname = generate_report(rtype, data)
        path = os.path.join(OUT, fname)
        with open(path, "wb") as f:
            f.write(buf.read())
        print(f"✓  {fname}  ({rtype}, dealers={len(data.get('dealers', [data]))})")
