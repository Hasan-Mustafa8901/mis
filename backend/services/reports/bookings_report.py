import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def booking_report_generator(df: pd.DataFrame, start_date, end_date):
    """
    Generate a formatted Excel booking report and return as BytesIO for Streamlit download.
    """
    # --- File Name ---
    if start_date == end_date:
        date = f"{start_date.strftime("%d-%m-%Y")}"
    else:
        date = f"{start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}"
    file_name = f"{date}-Report.xlsx"

    # --- Create Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # --- Title Row (merged) ---
    title = f"Booking Report - All Dealers"
    merge_end_col = get_column_letter(len(df.columns))
    ws.merge_cells(f"A1:{merge_end_col}1")
    title_cell = ws.cell(row=1,column=1)
    title_cell.value = title
    title_cell.font = Font(name="Times New Roman", bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Define border style
    thin_border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply style to all cells in merged range for proper border display
    start_col = 1
    end_col = len(df.columns)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    # ws.row_dimensions[1].height = 30

    # ---- Time Range (merged) ---
    ws.merge_cells(f"A2:{merge_end_col}2")
    date_cell = ws["A2"]
    date_cell.value = date
    date_cell.font = Font(name="Times New Roman",bold=True, color="000000")
    # title_cell.fill = PatternFill("solid", fgColor="247FD4")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    # ws.row_dimensions[2].height = 30
    idx_col = ws.cell(row=3,column=1)
    idx_col.value = "Date"
    idx_col.font = Font(name="Times New Roman",bold=True, color="000000")
    idx_col.fill = PatternFill("solid", fgColor="D3D3D3")
    idx_col.alignment = Alignment(horizontal="center", vertical="center")

    # --- Write column headers (row 2) ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(name="Times New Roman",bold=True, color="000000")
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    # ws.row_dimensions[2].height = 25

    # --- Write DataFrame rows (starting row 3) ---
    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx % 2==0:
                cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Safely auto-adjust column widths ---
    for col_idx in range(1, len(df.columns)):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    # --- Save to memory ---
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, file_name

def delivery_report_generator(df: pd.DataFrame, start_date, end_date):
    """
    Generate a formatted Excel booking report and return as BytesIO for Streamlit download.
    """

    # --- File Name ---
    if start_date == end_date:
        date = f"{start_date.strftime("%d-%m-%Y")}"
    else:
        date = f"{start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}"
    file_name = f"{date}-Report.xlsx"

    # --- Create Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries"

    # --- Title Row (merged) ---
    title = f"Delivery Report - All Dealers"
    merge_end_col = get_column_letter(len(df.columns))
    ws.merge_cells(f"A1:{merge_end_col}1")
    title_cell = ws.cell(row=1,column=1)
    title_cell.value = title
    title_cell.font = Font(name="Times New Roman", bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Define border style
    thin_border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply style to all cells in merged range for proper border display
    start_col = 1
    end_col = len(df.columns)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    # ws.row_dimensions[1].height = 30

    # ---- Time Range (merged) ---
    ws.merge_cells(f"A2:{merge_end_col}2")
    date_cell = ws["A2"]
    date_cell.value = date
    date_cell.font = Font(name="Times New Roman",bold=True, color="000000")
    # title_cell.fill = PatternFill("solid", fgColor="247FD4")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    # ws.row_dimensions[2].height = 30
    idx_col = ws.cell(row=3,column=1)
    idx_col.value = "Date"
    idx_col.font = Font(name="Times New Roman",bold=True, color="000000")
    idx_col.fill = PatternFill("solid", fgColor="D3D3D3")
    idx_col.alignment = Alignment(horizontal="center", vertical="center")

    # --- Write column headers (row 2) ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(name="Times New Roman",bold=True, color="000000")
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    # ws.row_dimensions[2].height = 25

    # --- Write DataFrame rows (starting row 3) ---
    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx % 2==0:
                cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Safely auto-adjust column widths ---
    for col_idx in range(1, len(df.columns)):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    # --- Save to memory ---
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, file_name