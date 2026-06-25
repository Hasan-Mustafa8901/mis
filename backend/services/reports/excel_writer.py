# backend\services\reports\excel_writer.py
from typing import Dict, Any
from sqlmodel import Session, select, col
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell

from db.models import DiscountComponent
from services.reports.export_query import query_export_transactions_batch
from datetime import date
from rich import print


def export_mis_excel_incremental(
    session: Session,
    file_path: str,
    query_args: Dict[str, Any],
    batch_size: int = 500,
) -> int:
    """
    Incrementally write MIS export chunks to Excel using Workbook(write_only=True) to avoid memory overhead.
    Returns the total row count exported.
    """
    # 1. Fetch Discount Components for dynamic columns
    components = session.exec(
        select(DiscountComponent).order_by(col(DiscountComponent.order))
    ).all()

    sample_batch = query_export_transactions_batch(
        session=session,
        **query_args,
        last_id=None,
        limit=100,
    )

    dynamic_json_headers = set()

    for row in sample_batch:
        for key in row.keys():
            if (
                key.startswith("invoice_")
                or key.startswith("payment_")
                or key.startswith("audit_")
                or key.startswith("booking_checklist_")
                or key.startswith("delivery_checklist_")
            ):
                dynamic_json_headers.add(key)

    dynamic_json_headers = sorted(dynamic_json_headers)

    # 2. Build ordered headers list
    headers = [
        "Export ID",
        "Status",
        "Stage",
        "Mode",
        "Created By",
        "Created At",
        "Showroom/Outlet",
        "Dealership",
        "Customer Name",
        "Mobile Number",
        "Alternate Mobile",
        "Email",
        "PAN Number",
        "Aadhar Number",
        "Address",
        "City",
        "Pin Code",
        "Car Name",
        "Variant Name",
        "Full Variant Name",
        "Fuel Type",
        "VIN Number",
        "Engine Number",
        "Color",
        "Registration Number",
        "Registration Date",
        "Model Year",
        "Booking Date",
        "Booking Amt",
        "Booking Receipt Num",
        "Delivery Date",
        "Invoice Number",
        "Customer File Number",
        "Sales Executive",
        "Team Leader",
        "Booking File Incomplete",
        "Booking File Incomplete Remarks",
        "Delivery File Incomplete",
        "Delivery File Incomplete Remarks",
        "Adjustment Delivery",
        "Adjustment Booking",
        "Net Receivable",
        "Total Received",
        "Balance Amount",
        "Other Discount (Delivery)",
        "Other Discount (Booking)",
        "Payment Status",
        # "Accessories",
        "Total Allowed Discount",
        "Total Actual Discount",
        "Total Excess Discount",
    ]
    # Add dynamic columns
    for comp in components:
        headers.append(f"{comp.name} Actual")
        headers.append(f"{comp.name} Allowed")

    # Add condition columns
    cond_keys = [
        "corporate",
        "exchange",
        "scrappage",
        "loyalty",
        "sbi_yono",
        "solar_roof_top",
    ]
    for ck in cond_keys:
        headers.append(f"Cond: {ck.replace('_', ' ').title()}")

    headers.extend(dynamic_json_headers)

    # 3. Create Write-Only Workbook & Styles
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="MIS Export")

    # Styling elements
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    header_font = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")  # Premium Navy Blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Times New Roman", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(
        "solid", fgColor="F9FAFB"
    )  # Subtle light grey for alternate rows
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    # Write Header row
    header_row = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        header_row.append(cell)
    ws.append(header_row)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Track column widths for formatting
    col_widths = {i: len(h) for i, h in enumerate(headers, 1)}

    last_id = None
    total_rows = 0

    while True:
        # Fetch chunk
        batch = query_export_transactions_batch(
            session=session, **query_args, last_id=last_id, limit=batch_size
        )

        if not batch:
            break

        for txn in batch:
            row_cells = []
            row_idx = total_rows + 1  # alternating index
            current_fill = alt_fill if row_idx % 2 == 0 else white_fill

            # Map fields in header order
            # Basic transaction details
            values = [
                txn.get("id"),
                txn.get("status"),
                txn.get("stage"),
                txn.get("mode"),
                txn.get("created_by_name") or "System",
                txn.get("created_at").strftime("%Y-%m-%d %H:%M:%S")
                if txn.get("created_at")
                else "",
                txn.get("outlet_name"),
                txn.get("dealership_name"),
                txn.get("customer_name"),
                txn.get("customer_mobile_number"),
                txn.get("customer_alternate_mobile"),
                txn.get("customer_email"),
                txn.get("customer_pan_number"),
                txn.get("customer_aadhar_number"),
                txn.get("customer_address"),
                txn.get("customer_city"),
                txn.get("customer_pin_code"),
                txn.get("car_name"),
                txn.get("variant_name"),
                txn.get("full_variant_name"),
                txn.get("fuel_type"),
                txn.get("vin_number"),
                txn.get("engine_number"),
                txn.get("color"),
                txn.get("registration_number"),
                txn.get("registration_date").isoformat()
                if isinstance(txn.get("registration_date"), date)
                else txn.get("registration_date"),
                txn.get("model_year"),
                txn.get("booking_date").isoformat()
                if isinstance(txn.get("booking_date"), date)
                else txn.get("booking_date"),
                txn.get("booking_amt"),
                txn.get("booking_receipt_num"),
                txn.get("delivery_date").isoformat()
                if isinstance(txn.get("delivery_date"), date)
                else txn.get("delivery_date"),
                txn.get("invoice_number"),
                txn.get("customer_file_number"),
                txn.get("sales_executive_name"),
                txn.get("team_leader"),
                "Yes" if txn.get("booking_file_incomplete") else "No",
                txn.get("booking_file_incomplete_remarks") or "",
                "Yes" if txn.get("delivery_file_incomplete") else "No",
                txn.get("delivery_file_incomplete_remarks") or "",
                txn.get("adjustment_delivery"),
                txn.get("adjustment_booking"),
                txn.get("total_receivable"),
                txn.get("total_received"),
                txn.get("balance"),
                txn.get("other_discount_delivery"),
                txn.get("other_discount_booking"),
                txn.get("payment_status") or "",
                # ", ".join(txn.get("accessories") or []),
                txn.get("total_allowed_discount"),
                txn.get("total_actual_discount"),
                txn.get("total_excess_discount"),
            ]

            # Dynamic components values
            items_data = txn.get("items") or {}
            for comp in components:
                comp_vals = items_data.get(comp.name, {"actual": 0.0, "allowed": 0.0})
                values.append(comp_vals.get("actual", 0.0))
                values.append(comp_vals.get("allowed", 0.0))

            # Condition values
            cond_data = txn.get("conditions") or {}
            for ck in cond_keys:
                val = cond_data.get(ck)
                values.append("Yes" if val else "No")

            for header in dynamic_json_headers:
                values.append(txn.get(header, ""))

            # Create Cell objects with styling
            for col_idx, val in enumerate(values, 1):
                # Clean value to avoid None representation issues
                cell_val = "" if val is None else val
                cell = WriteOnlyCell(ws, value=cell_val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = border
                cell.fill = current_fill
                row_cells.append(cell)

                if len(values) != len(headers):
                    print("HEADER COUNT:", len(headers))
                    print("VALUE COUNT :", len(values))
                    print("EXTRA COLS  :", len(values) - len(headers))
                    raise Exception(
                        f"Header/Value mismatch. headers={len(headers)} values={len(values)}"
                    )

                # Keep track of width
                val_len = len(str(cell_val))
                if val_len > col_widths[col_idx]:
                    col_widths[col_idx] = val_len

            ws.append(row_cells)
            total_rows += 1

        last_id = batch[-1]["id"]

    # 4. Apply column widths (enforced range [10, 35])
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(width + 3, 10), 35)

    wb.save(file_path)
    return total_rows
