from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── DATA: Add / remove / rename keys freely — columns auto-adjust ──────────────
DATA = [
    {
        "Select": "No",
        "S.No": 1,
        "Booking Date": "27/05/2026",
        "Customer Name": "CHANDRA SINGH",
        "Mobile": "7379732222",
        "Car Model": "Safari 2.0",
        "TL": "Pradeep Sharma",
        "ReceivingDate": "—",
        "Out of ScopeReason": "",
        "Approved": "No",
        "RejectionReason": "",
        "Scanned Date": "",
        "MIS Entry": "—",
        "Incomplete": "No",
        "Incomplete Remarks": "—",
        "Received": "",
    },
    {
        "Select": "No",
        "S.No": 2,
        "Booking Date": "27/05/2026",
        "Customer Name": "ISRAR AHMAD KHAN",
        "Mobile": "9140287012",
        "Car Model": "Punch",
        "TL": "Pramod Kumar Singh",
        "ReceivingDate": "29/05/2026",
        "Out of ScopeReason": "Booking before 1 may",
        "Approved": "No",
        "RejectionReason": "",
        "Scanned Date": "",
        "MIS Entry": "—",
        "Incomplete": "No",
        "Incomplete Remarks": "—",
        "Received": "Yes",
    },
    {
        "Select": "No",
        "S.No": 3,
        "Booking Date": "27/05/2026",
        "Customer Name": "JYOTI VERMA",
        "Mobile": "9451516000",
        "Car Model": "Nexon EV 3.0",
        "TL": "Gaurav Mishra",
        "ReceivingDate": "27/05/2026",
        "Out of ScopeReason": "",
        "Approved": "No",
        "RejectionReason": "",
        "Scanned Date": "27/05/2026",
        "MIS Entry": "—",
        "Incomplete": "No",
        "Incomplete Remarks": "—",
        "Received": "Yes",
    },
]

# ── EXPORT ─────────────────────────────────────────────────────────────────────
def export_mis_excel(data: list[dict], output_path: str) -> str:
    # Column headers come 100% from dict keys — nothing hardcoded
    headers = list(dict.fromkeys(key for row in data for key in row.keys()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Booking MIS"

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # Data rows
    for record in data:
        ws.append([record.get(h, "") for h in headers])
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 18
        for cell in ws[row_idx]:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

    # Auto column width based on content
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len    = max(
            len(str(header)),
            *[len(str(row.get(header, ""))) for row in data]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    filename = f"MIS_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"Done → {export_mis_excel(DATA, filename)}")